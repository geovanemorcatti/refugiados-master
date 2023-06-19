from pathlib import Path
import openpyxl
from datetime import datetime
from tempfile import NamedTemporaryFile
from django.http import HttpResponse
from core.models import Refugiado


def get_model_xls():
  current_dir = Path(__file__).resolve().parent
  return openpyxl.load_workbook(f'{current_dir}/model.xlsx')


def ternary_or_empty(payload: dict, key: str):
  if payload.get(key, None):
    return 'X'
  else:
    return '  '


def ternary_or_empty_reverse(payload: dict, key: str):
  if payload.get(key, None):
    return '  '
  else:
    return 'X'


def date_or_empty(payload: dict, key: str):
  try:
    dt = payload.get(key, None).strftime('%m/%d/%Y')
  except:
    dt = ''
  return dt


def to_date_or_none(obj):
  try:
    dt = obj.strftime('%Y-%m-%d')
  except:
    dt = None
  return dt


def export_to_xls(payload: dict, ref: Refugiado):
  xls = get_model_xls()
  sheet = xls.active

  print(payload)
  
  sheet['AM3'] = ref.id

  sheet['B7'] = payload.get('dataPAtendimento', '')
  sheet['H6'] = f'({ternary_or_empty(payload, "dgTraficoDePessoas")}) Tráfico de Pessoas'
  sheet['T6'] = f'({ternary_or_empty(payload, "dgTrabalhoEscravo")}) Trabalho Escravo'
  sheet['AE6'] = f'({ternary_or_empty(payload, "dgViolenciaDomestica")}) Violência doméstica'
  sheet['H7'] = f'({ternary_or_empty(payload, "dgRegulacaoMigratoria")}) Regularização Migratória'
  sheet['T7'] = f'({ternary_or_empty(payload, "dgServicoSocial")}) Serviço Social'
  sheet['AE7'] = f'({ternary_or_empty(payload, "dgJuridico")}) Jurídico'

  sheet['H9'] = ref.nome
  sheet['AJ9'] = payload.get('estadoCivil', '')
  sheet['H10'] = date_or_empty(payload, 'dataNascimento')
  sheet['Q10'] = ref.idade
  sheet['W10'] = payload.get('religiao', '')
  sheet['H11'] = payload.get('nomeDaMae', '')
  sheet['AE11'] = payload.get('grauDeInstrucao', '')
  sheet['H12'] = payload.get('nomeDoPai', '')
  sheet['H13'] = payload.get('paisNascimento', '')
  sheet['P13'] = payload.get('cidadeNascimento', '')
  sheet['W13'] = ref.sexo
  sheet['E14'] = ref.lingua
  sheet['G15'] = payload.get('dni', '')
  sheet['Q15'] = payload.get('passaporte', '')
  sheet['Z15'] = payload.get('rne', '')
  sheet['AI15'] = payload.get('cpf', '')
  sheet['E16'] = ref.email
  sheet['W16'] = payload.get('telefone', '')
  sheet['AI16'] = ref.celular
  sheet['I17'] = ref.motivo_ida_brasil
  sheet['AK17'] = payload.get('dataIngresso', '')
  sheet['J18'] = payload.get('rotaChegadaBrasil', '')
  sheet['AN18'] = payload.get('numeroDeFilhos', '')
  
  sheet['H20'] = payload.get('enderecoBrasil', '')
  sheet['AH20'] = payload.get('enderecoBrasilBairro', '')
  sheet['E21'] = payload.get('enderecoBrasilCidade', '')
  sheet['M21'] = payload.get('enderecoBrasilEstado', '')
  sheet['X21'] = payload.get('enderecoBrasilCEP', '')
  sheet['AJ21'] = payload.get('enderecoBrasilComplemento', '')
  sheet['N22'] = payload.get('enderecoFamilia', '')
  
  sheet['G24'] = payload.get('profissaoAtual', '')
  sheet['S24'] = f'({ternary_or_empty_reverse(payload, "carteiraDeTrabalho")}) Não'
  sheet['V24'] = f'({ternary_or_empty(payload, "carteiraDeTrabalho")}) Sim'
  sheet['AB24'] = payload.get('carteiraDeTrabalho', '')
  sheet['G25'] = payload.get('cargoOuFuncao', '')
  sheet['AE25'] = payload.get('outrasInfos', '')
  sheet['E26'] = payload.get('empresa', '')
  sheet['P26'] = payload.get('localDeTrabalho', '')
  sheet['AJ26'] = payload.get('horarioTrabalho', '')
  
  sheet['K28'] = payload.get('problemasSaude', '')
  sheet['G29'] = payload.get('medicamentos', '')
  sheet['G29'] = payload.get('medicamentos', '')
  sheet['F30'] = f'({ternary_or_empty_reverse(payload, "cartaoSus")}) Não'
  sheet['K30'] = payload.get('cartaoSus', '')
  sheet['T30'] = f'({ternary_or_empty_reverse(payload, "beneficios")}) Não'
  sheet['V30'] = payload.get('beneficios', '')
  sheet['AG28'] = payload.get('saudeVicios', '')

  with NamedTemporaryFile() as tmp:
    xls.save(tmp.name)
    tmp.seek(0)
    stream = tmp.read()

  response = HttpResponse(content=stream, content_type='application/ms-excel', )
  response['Content-Disposition'] = f'attachment; filename=Refugiado-{ref.nome}-export.xlsx'
  return response


def import_from_xls(file):
  book = openpyxl.load_workbook(file)
  sheet = book.active

  create_count = 0
  for row in sheet.iter_rows(min_row=4, min_col=1, max_col=58, values_only=True):
    print(row)
    col_list = list(row)
    if len(str(col_list[0])) > 0 and not Refugiado.objects.filter(nome=col_list[0]).exists():
      new_ref = Refugiado(
        nome = col_list[0],
        sexo = col_list[1],
        idade = col_list[2],
        nacionalidade = col_list[3],
        formacao = col_list[4],
        profissao = col_list[5],
        lingua = col_list[6],
        pais = col_list[7],
        cidade = col_list[8],
        primeiro_destino = col_list[9],
        motivo_ida_brasil = col_list[10],
        como_conheceu_projeto = col_list[11],
        servicos_que_procura = col_list[12],
        celular = col_list[35],
        email = col_list[33],
        ficha_social = {
          'dataPAtendimento': to_date_or_none(col_list[14]),
          'dgTraficoDePessoas': len(str(col_list[15])) > 0,
          'dgTrabalhoEscravo': len(str(col_list[16])) > 0,
          'dgViolenciaDomestica': len(str(col_list[17])) > 0,
          'dgRegulacaoMigratoria': len(str(col_list[18])) > 0,
          'dgServicoSocial': len(str(col_list[19])) > 0,
          'dgJuridico': len(str(col_list[20])) > 0,
          'estadoCivil': col_list[21],
          'dataNascimento': to_date_or_none(col_list[22]),
          'religiao': col_list[23],
          'nomeDaMae': col_list[24],
          'nomeDoPai': col_list[25],
          'grauDeInstrucao': col_list[26],
          'paisNascimento': col_list[27],
          'cidadeNascimento': col_list[28],
          'dni': col_list[29],
          'rne': col_list[30],
          'passaporte': col_list[31],
          'cpf': col_list[32],
          'telefone': col_list[34],
          'dataIngresso': to_date_or_none(col_list[36]),
          'rotaChegadaBrasil': col_list[37],
          'numeroDeFilhos': col_list[38],
          'enderecoBrasil': col_list[39],
          'enderecoBrasilBairro': col_list[40],
          'enderecoBrasilCidade': col_list[41],
          'enderecoBrasilEstado': col_list[42],
          'enderecoBrasilCEP': col_list[43],
          'enderecoBrasilComplemento': col_list[44],
          'enderecoFamilia': col_list[45],
          'profissaoAtual': col_list[46],
          'carteiraDeTrabalho': col_list[47],
          'cargoOuFuncao': col_list[48],
          'outrasInfos': col_list[49],
          'empresa': col_list[50],
          'localDeTrabalho': col_list[51],
          'horarioTrabalho': col_list[52],
          'problemasSaude': col_list[53],
          'medicamentos': col_list[54],
          'cartaoSus': col_list[55],
          'beneficios': col_list[56],
          'saudeVicios': col_list[57],
        }
      )
      new_ref.save()
      create_count += 1
  
  return create_count

